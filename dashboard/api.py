"""Dashboard API for Dev Companion — findings, observability, settings, reports."""

import io
import json
import os
import datetime
from typing import Optional

import aiosqlite
from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles

from dashboard.db import DB_PATH, DEFAULT_SYSTEM_PROMPT, get_config, save_config, update_finding_status

_HERE = os.path.dirname(os.path.abspath(__file__))

dashboard_app = FastAPI(title="Dev Companion Dashboard")

try:
    dashboard_app.mount("/dashboard/static", StaticFiles(directory=os.path.join(_HERE, "static"), html=True), name="static")
except Exception:
    pass


@dashboard_app.get("/dashboard/", include_in_schema=False)
async def redirect():
    return RedirectResponse(url="/dashboard")


@dashboard_app.get("/dashboard", include_in_schema=False)
async def index():
    return FileResponse(os.path.join(_HERE, "static", "index.html"))


def _dates(start, end):
    today = datetime.date.today()
    return start or (today - datetime.timedelta(days=30)).isoformat(), end or today.isoformat()


# --- Config ---

@dashboard_app.get("/dashboard/api/config")
async def api_get_config():
    return await get_config()


@dashboard_app.put("/dashboard/api/config")
async def api_put_config(body: dict):
    if not isinstance(body.get("iteration_limit"), int) or not (1 <= body["iteration_limit"] <= 500):
        raise HTTPException(422, "iteration_limit must be 1-500")
    if not isinstance(body.get("system_prompt"), str) or not body["system_prompt"].strip():
        raise HTTPException(422, "system_prompt must be non-empty")
    if not isinstance(body.get("enabled_tools"), list):
        raise HTTPException(422, "enabled_tools must be a list")
    existing = await get_config()
    existing.update(body)
    await save_config(existing)
    return {"status": "ok"}


@dashboard_app.post("/dashboard/api/config/reset-prompt")
async def api_reset_prompt():
    cfg = await get_config()
    cfg["system_prompt"] = DEFAULT_SYSTEM_PROMPT
    await save_config(cfg)
    return {"status": "ok"}


# --- Workspaces ---

@dashboard_app.get("/dashboard/api/workspaces")
async def api_workspaces():
    """Return distinct project names that have findings recorded."""
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(
            "SELECT DISTINCT workspace FROM review_findings WHERE workspace IS NOT NULL AND workspace != '' ORDER BY workspace"
        ) as cur:
            rows = await cur.fetchall()
    return [r[0] for r in rows]


# --- Findings (code-reviewer specific) ---

@dashboard_app.get("/dashboard/api/findings")
async def api_findings(
    thread_id: Optional[str] = None,
    criticality: Optional[str] = None,
    category: Optional[str] = None,
    workspace: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
):
    query = "SELECT * FROM review_findings WHERE 1=1"
    params = []
    if start and end:
        query += " AND substr(timestamp,1,10) BETWEEN ? AND ?"
        params += [start, end]
    if thread_id:
        query += " AND thread_id = ?"
        params.append(thread_id)
    if criticality:
        query += " AND criticality = ?"
        params.append(criticality)
    if category:
        query += " AND category = ?"
        params.append(category)
    if workspace:
        query += " AND workspace = ?"
        params.append(workspace)
    query += " ORDER BY CASE criticality WHEN 'critical' THEN 1 WHEN 'high' THEN 2 WHEN 'medium' THEN 3 WHEN 'low' THEN 4 ELSE 5 END, timestamp DESC"

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(query, params) as cur:
            rows = [dict(r) for r in await cur.fetchall()]
    return rows


@dashboard_app.get("/dashboard/api/findings/summary")
async def api_findings_summary(workspace: Optional[str] = None, start: Optional[str] = None, end: Optional[str] = None):
    ws_clause = " WHERE workspace = ?" if workspace else ""
    ws_params = [workspace] if workspace else []
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(
            f"SELECT criticality, COUNT(*) as count FROM review_findings{ws_clause} GROUP BY criticality",
            ws_params) as cur:
            by_criticality = {r[0]: r[1] for r in await cur.fetchall()}

        async with db.execute(
            f"SELECT category, COUNT(*) as count FROM review_findings{ws_clause} GROUP BY category ORDER BY count DESC",
            ws_params) as cur:
            by_category = [{"category": r[0], "count": r[1]} for r in await cur.fetchall()]

        async with db.execute(f"SELECT COUNT(*) FROM review_findings{ws_clause}", ws_params) as cur:
            total = (await cur.fetchone())[0]

    return {
        "total": total,
        "by_criticality": by_criticality,
        "by_category": by_category,
    }


# --- Observability (same as ai-intern) ---

@dashboard_app.get("/dashboard/api/telemetry/summary")
async def api_telemetry_summary(start: Optional[str] = None, end: Optional[str] = None):
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(
            "SELECT COALESCE(SUM(prompt_tokens),0), COALESCE(SUM(completion_tokens),0), COALESCE(SUM(total_tokens),0), COUNT(*) FROM llm_calls") as cur:
            pt, ct, tt, calls = await cur.fetchone()
    return {"prompt_tokens": pt, "completion_tokens": ct, "total_tokens": tt, "llm_call_count": calls}


@dashboard_app.get("/dashboard/api/telemetry/tokens-over-time")
async def api_tokens_over_time(start: Optional[str] = None, end: Optional[str] = None):
    start, end = _dates(start, end)
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(
            """SELECT substr(timestamp,1,10) AS date,
               COALESCE(SUM(prompt_tokens),0), COALESCE(SUM(completion_tokens),0),
               COALESCE(SUM(total_tokens),0), COUNT(*)
               FROM llm_calls WHERE substr(timestamp,1,10) BETWEEN ? AND ?
               GROUP BY date ORDER BY date""", (start, end)) as cur:
            rows = await cur.fetchall()
    return [{"date": r[0], "prompt_tokens": r[1], "completion_tokens": r[2], "total_tokens": r[3], "call_count": r[4]} for r in rows]


@dashboard_app.get("/dashboard/api/telemetry/tools")
async def api_telemetry_tools(start: Optional[str] = None, end: Optional[str] = None):
    start, end = _dates(start, end)
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(
            """SELECT tool_name, COUNT(*) AS n,
               AVG(CASE WHEN duration_ms IS NOT NULL THEN duration_ms END),
               CAST(SUM(CASE WHEN status='failure' THEN 1 ELSE 0 END) AS REAL)/COUNT(*)*100
               FROM tool_invocations WHERE substr(timestamp,1,10) BETWEEN ? AND ?
               GROUP BY tool_name ORDER BY n DESC""", (start, end)) as cur:
            rows = await cur.fetchall()
    return [{"tool_name": r[0], "invocation_count": r[1], "avg_duration_ms": r[2] or 0.0, "failure_rate": r[3] or 0.0} for r in rows]


@dashboard_app.get("/dashboard/api/telemetry/sessions")
async def api_sessions(start: Optional[str] = None, end: Optional[str] = None):
    """Returns sessions from llm_calls for observability tab session count."""
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(
            "SELECT thread_id, MAX(timestamp) FROM llm_calls GROUP BY thread_id") as cur:
            sessions = {r[0]: r[1] for r in await cur.fetchall()}
    return sorted([{"thread_id": k, "last_active": v} for k, v in sessions.items()],
                  key=lambda x: x["last_active"], reverse=True)


@dashboard_app.get("/dashboard/api/export")
async def api_export():
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        llm = [dict(r) for r in await (await db.execute("SELECT * FROM llm_calls ORDER BY timestamp")).fetchall()]
        tools = [dict(r) for r in await (await db.execute("SELECT * FROM tool_invocations ORDER BY timestamp")).fetchall()]
        findings = [dict(r) for r in await (await db.execute("SELECT * FROM review_findings ORDER BY timestamp")).fetchall()]
    return Response(
        content=json.dumps({"llm_calls": llm, "tool_invocations": tools, "review_findings": findings}, indent=2, default=str),
        media_type="application/json",
        headers={"Content-Disposition": 'attachment; filename="review_export.json"'})


# --- New v2 endpoints ---

@dashboard_app.get("/dashboard/api/findings/by-file")
async def api_findings_by_file(thread_id: Optional[str] = None, workspace: Optional[str] = None):
    conditions = []
    params = []
    if thread_id:
        conditions.append("thread_id = ?")
        params.append(thread_id)
    if workspace:
        conditions.append("workspace = ?")
        params.append(workspace)
    where = (" WHERE " + " AND ".join(conditions)) if conditions else ""
    query = f"""SELECT file_path, COUNT(*) as total,
               SUM(CASE WHEN criticality='critical' THEN 1 ELSE 0 END) as critical,
               SUM(CASE WHEN criticality='high' THEN 1 ELSE 0 END) as high,
               SUM(CASE WHEN criticality='medium' THEN 1 ELSE 0 END) as medium,
               SUM(CASE WHEN criticality='low' THEN 1 ELSE 0 END) as low
               FROM review_findings{where}
               GROUP BY file_path ORDER BY total DESC LIMIT 10"""
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(query, params) as cur:
            return [dict(r) for r in await cur.fetchall()]


@dashboard_app.get("/dashboard/api/findings/trend")
async def api_findings_trend(workspace: Optional[str] = None, start: Optional[str] = None, end: Optional[str] = None):
    start, end = _dates(start, end)
    ws_clause = " AND workspace = ?" if workspace else ""
    ws_params = [workspace] if workspace else []
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(
            f"""SELECT substr(timestamp,1,10) as date, COUNT(*) as count
               FROM review_findings WHERE substr(timestamp,1,10) BETWEEN ? AND ?{ws_clause}
               GROUP BY date ORDER BY date""", [start, end] + ws_params) as cur:
            rows = await cur.fetchall()
    return [{"date": r[0], "count": r[1]} for r in rows]


@dashboard_app.get("/dashboard/api/findings/heatmap")
async def api_findings_heatmap(thread_id: Optional[str] = None, workspace: Optional[str] = None):
    conditions = []
    params = []
    if thread_id:
        conditions.append("thread_id = ?")
        params.append(thread_id)
    if workspace:
        conditions.append("workspace = ?")
        params.append(workspace)
    where = (" WHERE " + " AND ".join(conditions)) if conditions else ""
    query = f"SELECT category, criticality, COUNT(*) as count FROM review_findings{where} GROUP BY category, criticality"
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(query, params) as cur:
            rows = await cur.fetchall()
    result = {}
    for cat, crit, count in rows:
        if cat not in result:
            result[cat] = {}
        result[cat][crit] = count
    return result


@dashboard_app.get("/dashboard/api/sessions")
async def api_review_sessions():
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        try:
            async with db.execute(
                "SELECT * FROM review_sessions ORDER BY timestamp DESC") as cur:
                return [dict(r) for r in await cur.fetchall()]
        except Exception:
            return []


@dashboard_app.get("/dashboard/api/telemetry/sessions/{thread_id}")
async def api_session_detail(thread_id: str):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT * FROM llm_calls WHERE thread_id=? ORDER BY timestamp", (thread_id,)) as cur:
            llm_calls = [dict(r) for r in await cur.fetchall()]
        async with db.execute(
            "SELECT * FROM tool_invocations WHERE thread_id=? ORDER BY timestamp", (thread_id,)) as cur:
            tool_invs = [dict(r) for r in await cur.fetchall()]
        async with db.execute(
            "SELECT * FROM review_findings WHERE thread_id=? ORDER BY id", (thread_id,)) as cur:
            findings = [dict(r) for r in await cur.fetchall()]

    total_tokens = sum(r.get("total_tokens", 0) or 0 for r in llm_calls)
    return {
        "thread_id": thread_id,
        "llm_calls": llm_calls,
        "tool_invocations": tool_invs,
        "findings": findings,
        "total_tokens": total_tokens,
        "finding_count": len(findings),
    }


@dashboard_app.get("/dashboard/api/telemetry/models")
async def api_models(start: Optional[str] = None, end: Optional[str] = None):
    start, end = _dates(start, end)
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(
            """SELECT model, COUNT(*) as calls, SUM(total_tokens) as tokens
               FROM llm_calls WHERE substr(timestamp,1,10) BETWEEN ? AND ?
               GROUP BY model ORDER BY calls DESC""", (start, end)) as cur:
            rows = await cur.fetchall()
    return [{"model": r[0], "calls": r[1], "tokens": r[2] or 0} for r in rows]


@dashboard_app.get("/dashboard/api/telemetry/efficiency")
async def api_efficiency():
    """Tokens per finding per session — for the efficiency line chart."""
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(
            """SELECT l.thread_id, substr(MAX(l.timestamp),1,10) as date,
               SUM(l.total_tokens) as tokens, COUNT(f.id) as findings
               FROM llm_calls l
               LEFT JOIN review_findings f ON f.thread_id = l.thread_id
               GROUP BY l.thread_id
               HAVING tokens > 0
               ORDER BY date""") as cur:
            rows = await cur.fetchall()
    return [
        {"thread_id": r[0], "date": r[1],
         "tokens": r[2] or 0, "findings": r[3] or 0,
         "tokens_per_finding": round((r[2] or 0) / max(r[3], 1))}
        for r in rows
    ]


@dashboard_app.patch("/dashboard/api/findings/{finding_id}/status")
async def api_update_finding_status(finding_id: int, body: dict):
    status = body.get("status", "")
    if status not in ("open", "fixed", "dismissed"):
        raise HTTPException(422, "status must be open, fixed, or dismissed")
    await update_finding_status(finding_id, status)
    return {"status": "ok"}


# --- Report generation ---

def _criticality_color(crit: str) -> str:
    return {"critical": "#ef4444", "high": "#f97316", "medium": "#eab308",
            "low": "#3b82f6", "info": "#64748b"}.get(crit, "#64748b")


def _health_score(findings: list) -> float:
    weights = {"critical": 3.0, "high": 2.0, "medium": 1.0, "low": 0.3, "info": 0.1}
    penalty = sum(weights.get(f.get("criticality", "info"), 0.1) for f in findings)
    return max(0.0, round(10.0 - min(penalty, 10.0), 1))


@dashboard_app.get("/dashboard/api/reports/html")
async def api_report_html(thread_id: Optional[str] = None, workspace: Optional[str] = None):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        conditions = []
        params = []
        if thread_id:
            conditions.append("thread_id=?")
            params.append(thread_id)
        if workspace:
            conditions.append("workspace=?")
            params.append(workspace)
        where = (" WHERE " + " AND ".join(conditions)) if conditions else ""
        q = f"SELECT * FROM review_findings{where} ORDER BY CASE criticality WHEN 'critical' THEN 1 WHEN 'high' THEN 2 WHEN 'medium' THEN 3 WHEN 'low' THEN 4 ELSE 5 END"
        async with db.execute(q, params) as cur:
            findings = [dict(r) for r in await cur.fetchall()]

    if not findings:
        raise HTTPException(404, "No findings found for this session")

    date_str = datetime.date.today().isoformat()
    score = _health_score(findings)
    counts = {}
    for f in findings:
        counts[f["criticality"]] = counts.get(f["criticality"], 0) + 1

    # Group by file
    by_file = {}
    for f in findings:
        fp = f.get("file_path", "unknown")
        by_file.setdefault(fp, []).append(f)

    def badge(crit):
        color = _criticality_color(crit)
        return f'<span style="background:{color};color:#fff;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600">{crit.upper()}</span>'

    summary_cards = "".join(
        f'<div style="background:{_criticality_color(c)};color:#fff;padding:16px 24px;border-radius:8px;text-align:center">'
        f'<div style="font-size:28px;font-weight:700">{counts.get(c,0)}</div>'
        f'<div style="font-size:12px;margin-top:4px">{c.upper()}</div></div>'
        for c in ["critical", "high", "medium", "low", "info"]
    )

    findings_rows = "".join(
        f"<tr><td>{i+1}</td><td>{badge(f['criticality'])}</td>"
        f"<td><span style='color:#a5b4fc;font-family:monospace'>{f.get('file_path','')}"
        f"{':%d'%f['line_number'] if f.get('line_number') else ''}</span></td>"
        f"<td>{f.get('category','')}</td>"
        f"<td style='font-weight:500'>{f.get('title','')}</td>"
        f"<td style='color:#94a3b8;font-size:12px'>{f.get('description','')}</td>"
        f"<td style='color:#6ee7b7;font-size:12px'>{f.get('suggestion','')}</td></tr>"
        for i, f in enumerate(findings)
    )

    file_sections = ""
    for fp, flist in by_file.items():
        rows = "".join(
            f"<tr><td>{badge(f['criticality'])}</td>"
            f"<td>{'L%d'%f['line_number'] if f.get('line_number') else '—'}</td>"
            f"<td style='font-weight:500'>{f.get('title','')}</td>"
            f"<td style='color:#94a3b8;font-size:12px'>{f.get('description','')}</td>"
            f"<td style='color:#6ee7b7;font-size:12px'>{f.get('suggestion','')}</td></tr>"
            for f in flist
        )
        file_sections += (
            f"<details style='margin-bottom:12px'><summary style='cursor:pointer;padding:10px;background:#1a1d2e;"
            f"border-radius:6px;font-weight:600'>{fp} ({len(flist)} findings)</summary>"
            f"<table style='width:100%;margin-top:8px'><thead><tr>"
            f"<th>Crit</th><th>Line</th><th>Title</th><th>Description</th><th>Suggestion</th>"
            f"</tr></thead><tbody>{rows}</tbody></table></details>"
        )

    html = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"/>
<title>Code Review Report — {date_str}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0f1117;color:#e2e8f0;padding:32px}}
h1{{font-size:24px;font-weight:700;margin-bottom:4px}}
h2{{font-size:16px;font-weight:600;margin:32px 0 12px;color:#94a3b8;text-transform:uppercase;letter-spacing:.05em}}
table{{width:100%;border-collapse:collapse;font-size:13px;margin-bottom:24px}}
th{{text-align:left;padding:8px 12px;color:#64748b;font-size:11px;text-transform:uppercase;border-bottom:1px solid #2d3748}}
td{{padding:8px 12px;border-bottom:1px solid #1e2433;vertical-align:top}}
tr:hover td{{background:#1a1d2e}}
input{{background:#1a1d2e;border:1px solid #2d3748;color:#e2e8f0;padding:6px 10px;border-radius:6px;font-size:13px;width:300px}}
</style>
</head><body>
<h1>🔍 Code Review Report</h1>
<p style="color:#64748b;margin-top:4px">Generated {date_str} &nbsp;·&nbsp; {len(findings)} findings &nbsp;·&nbsp; Health score: <strong style="color:#6ee7b7">{score}/10</strong></p>

<h2>Executive Summary</h2>
<div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:24px">{summary_cards}</div>

<h2>All Findings</h2>
<input type="text" id="search" placeholder="Filter findings..." oninput="filterTable()" style="margin-bottom:12px"/>
<table id="findings-table">
<thead><tr><th>#</th><th>Criticality</th><th>File</th><th>Category</th><th>Title</th><th>Description</th><th>Suggestion</th></tr></thead>
<tbody>{findings_rows}</tbody>
</table>

<h2>By File</h2>
{file_sections}

<script>
function filterTable(){{
  const q=document.getElementById('search').value.toLowerCase();
  document.querySelectorAll('#findings-table tbody tr').forEach(r=>{{
    r.style.display=r.textContent.toLowerCase().includes(q)?'':'none';
  }});
}}
</script>
</body></html>"""

    return Response(
        content=html,
        media_type="text/html",
        headers={"Content-Disposition": f'attachment; filename="review_{date_str}.html"'})


@dashboard_app.get("/dashboard/api/reports/xlsx")
async def api_report_xlsx(thread_id: Optional[str] = None, workspace: Optional[str] = None):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        conditions = []
        params = []
        if thread_id:
            conditions.append("thread_id=?")
            params.append(thread_id)
        if workspace:
            conditions.append("workspace=?")
            params.append(workspace)
        where = (" WHERE " + " AND ".join(conditions)) if conditions else ""
        q = f"SELECT * FROM review_findings{where} ORDER BY CASE criticality WHEN 'critical' THEN 1 WHEN 'high' THEN 2 WHEN 'medium' THEN 3 WHEN 'low' THEN 4 ELSE 5 END"
        async with db.execute(q, params) as cur:
            findings = [dict(r) for r in await cur.fetchall()]

        llm_conditions = ["1=1"]
        llm_params = []
        if thread_id:
            llm_conditions.append("thread_id=?")
            llm_params.append(thread_id)
        async with db.execute(
            f"SELECT * FROM llm_calls WHERE {' AND '.join(llm_conditions)} ORDER BY timestamp",
            llm_params) as cur:
            llm_calls = [dict(r) for r in await cur.fetchall()]

    if not findings:
        raise HTTPException(404, "No findings found")

    wb = openpyxl.Workbook()
    crit_fills = {
        "critical": PatternFill("solid", fgColor="7F1D1D"),
        "high":     PatternFill("solid", fgColor="7C2D12"),
        "medium":   PatternFill("solid", fgColor="713F12"),
        "low":      PatternFill("solid", fgColor="1E3A5F"),
        "info":     PatternFill("solid", fgColor="1E293B"),
    }
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="312E81")

    def _header(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    counts = {}
    for f in findings:
        counts[f["criticality"]] = counts.get(f["criticality"], 0) + 1
    _header(ws, ["Criticality", "Count"])
    for crit in ["critical", "high", "medium", "low", "info"]:
        ws.append([crit.upper(), counts.get(crit, 0)])
    ws.append([])
    ws.append(["Code Health Score", _health_score(findings)])
    ws.append(["Total Findings", len(findings)])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15

    # Sheet 2: All Findings
    ws2 = wb.create_sheet("All Findings")
    cols = ["ID", "File", "Line", "Criticality", "Category", "Title", "Description", "Suggestion", "Status"]
    _header(ws2, cols)
    for f in findings:
        row = ws2.append([
            f.get("id"), f.get("file_path"), f.get("line_number"),
            f.get("criticality", "").upper(), f.get("category"),
            f.get("title"), f.get("description"), f.get("suggestion"),
            f.get("status", "open"),
        ])
        crit = f.get("criticality", "info")
        fill = crit_fills.get(crit)
        if fill:
            ws2.cell(ws2.max_row, 4).fill = fill
            ws2.cell(ws2.max_row, 4).font = Font(color="FFFFFF", bold=True)
    for i, w in enumerate([8, 30, 8, 12, 15, 30, 50, 50, 10], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Sheet 3: By File
    ws3 = wb.create_sheet("By File")
    _header(ws3, ["File", "Total", "Critical", "High", "Medium", "Low"])
    by_file = {}
    for f in findings:
        fp = f.get("file_path", "unknown")
        by_file.setdefault(fp, []).append(f)
    for fp, flist in sorted(by_file.items(), key=lambda x: -len(x[1])):
        fc = {}
        for f in flist:
            fc[f["criticality"]] = fc.get(f["criticality"], 0) + 1
        ws3.append([fp, len(flist), fc.get("critical",0), fc.get("high",0), fc.get("medium",0), fc.get("low",0)])
    ws3.column_dimensions["A"].width = 40

    # Sheet 4: Observability
    ws4 = wb.create_sheet("Observability")
    _header(ws4, ["Timestamp", "Model", "Prompt Tokens", "Completion Tokens", "Total Tokens"])
    for r in llm_calls:
        ws4.append([r.get("timestamp"), r.get("model"), r.get("prompt_tokens",0),
                    r.get("completion_tokens",0), r.get("total_tokens",0)])
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.date.today().isoformat()
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="review_{date_str}.xlsx"'})
